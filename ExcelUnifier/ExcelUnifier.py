from os import path
from os import scandir
from openpyxl import load_workbook
from Service import Service
import ServicesToExcel
import sys

OK = 0
FILE_LOADING_ERROR = 1
FILE_FORMAT_ERROR = 2
FILE_ACCESS_PERMISSION_ERROR = 3

def main():
    """Loads and modifies a sample spreadsheet in Excel."""
    flags = {}
    flag = None
    for arg in sys.argv:
        if '--' in arg:
            flag = arg
        else:
            if flag:
                flags[flag] = arg
                flag = None

    
    if '--p' in flags:
        root_path = flags['--p']
        print('Using custom path')
    else:
        root_path = path.dirname(__file__)

    if '--o' in flags:
        output_path = flags['--o']
    else:
        output_path = 'Servicios.xlsx'
       
    total_services = process_directory(root_path)
    unified_workbook = ServicesToExcel.create_unified_workbook()
    unified_workbook = ServicesToExcel.convert_to_excel(unified_workbook, 2, total_services)
    print('Total services created: ' + str(len(total_services)) + ' services')
    try:
        unified_workbook.save(output_path)
        print('Excel file created successfully')
    except PermissionError:
        print('Program do not have permission to access ' + output_path)
        sys.exit(FILE_ACCESS_PERMISSION_ERROR)
    print('Job Done...')
    

def process_directory(dir):    
    services_by_directory = []
    for entry in scandir(dir):
        if entry.is_dir():
            services_by_directory = services_by_directory + process_directory(entry.path)
        else:
            print('processing entry: ' + entry.name)
            temp_wb = load_workbook_from_path(entry.path)
            services_by_directory = services_by_directory + process_workbook(temp_wb, entry.name)
            print('processing... done')
    if len(services_by_directory) <= 0:
        print(dir + ' could not load any services')
    return services_by_directory


def load_workbook_from_path(workbook_path):
    try:
        workbook = load_workbook(workbook_path, read_only=True)
        print('The file has been loaded successfully')
    except FileNotFoundError:
        print('The file has not been found in ' + workbook_path);
        sys.exit(FILE_LOADING_ERROR)
    except FileExistsError:
        print('The file ' + workbook_path + ' doesn\'t exist')
        sys.exit(FILE_LOADING_ERROR)
    return workbook


def process_workbook(wb, workbook_path):
    sheet_names = wb.sheetnames
    servicesToParse = []
    for sheet in sheet_names:
        print('Getting services of sheet ' + sheet)
        servicesToParse = servicesToParse + process_sheet(sheet, wb)

    print(str(len(servicesToParse)) + ' services has been created from file ' + workbook_path)
    return servicesToParse


def process_sheet(sheet, workbook):   
    services = []    
    sheet_ranges = workbook[sheet]    
    columnDict = get_columns_index(sheet_ranges)

    max_col = sheet_ranges.rows.gi_frame.f_locals['max_col']
    max_row = sheet_ranges.rows.gi_frame.f_locals['max_row']

    STARTING_SERVICE_ROW = 2

    for row in sheet_ranges.iter_rows(min_row=STARTING_SERVICE_ROW, max_col=max_col, max_row=max_row):
        fields = {}
        for cell in row: 
            if hasattr(cell,'column') and  cell.column in columnDict.keys():
                fields[columnDict[cell.column]] = cell.value
            else:
                continue
        
        if len(fields) == 0:
            continue

        service = Service(fields)
        if service.is_valid_service():
            services.append(service)
        
    print('Services for ' + sheet + ' has been processed')   
    return services
    
           

def get_columns_index(sheet_ranges):
    columnIdx = {}
    max_col = sheet_ranges.rows.gi_frame.f_locals['max_col']    
    for j in range(1,(max_col+1)):
        cell = sheet_ranges.cell(row=1, column=j)
        columnName = get_dict_equivalent(cell.value)
        if columnName:
            columnIdx[j] = columnName
    return columnIdx


def get_dict_equivalent(columnName):
    dict = {
        'FECHA':'date',
        'AEROLINEA':'airline',
        'ORIGEN':'origin',
        'DESTINO':'destination',
        'NO VUELO':'flight_number',
        'NOMBRE PASAJERO':'pax_name',
        'TERMINAL':'service_terminal',
        'AREA':'service_terminal',
        'N° VUELO':'flight_number',
        'WCOB':'WCOB',
        'WCMP':'WCMP',
        'WCHS':'WCHS',
        'WCHR':'WCHR',
        'WCHC':'WCHC',
        'WCBW':'WCBW'
    }
    if columnName in dict.keys():
        return dict[columnName]
    else:
        return None   


if __name__ == '__main__':
    main()
