from openpyxl import Workbook


def create_unified_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Servicios Avianca'
    ws['A1'] = 'FECHA'
    ws['B1'] = 'TERMINAL'
    ws['C1'] = 'NOMBRE PASAJERO'
    ws['D1'] = 'NO VUELO'
    ws['E1'] = 'ORIGEN'
    ws['F1'] = 'DESTINO'
    ws['G1'] = 'AEROLINEA'
    ws['H1'] = 'WCOB'
    ws['I1'] = 'WCMP'
    ws['J1'] = 'WCHS'
    ws['K1'] = 'WCHR'
    ws['L1'] = 'WCHC'
    ws['M1'] = 'WCBW'
    return wb

def convert_to_excel(wb, starting_row, services = []):
    ws = wb['Servicios Avianca']
    total_services = len(services)
    for i in range(0,total_services):
        offset = i + starting_row
        ws['A' + str(offset)] = services[i].date
        ws['B' + str(offset)] = services[i].service_terminal
        ws['C' + str(offset)] = services[i].pax_name
        ws['D' + str(offset)] = services[i].flight_number
        ws['E' + str(offset)] = services[i].origin
        ws['F' + str(offset)] = services[i].destination
        ws['G' + str(offset)] = services[i].airline
        ws['H' + str(offset)] = services[i].wcob
        ws['I' + str(offset)] = services[i].wcmp
        ws['J' + str(offset)] = services[i].wchs
        ws['K' + str(offset)] = services[i].wchr
        ws['L' + str(offset)] = services[i].wchc
        ws['M' + str(offset)] = services[i].wcbw

    return wb